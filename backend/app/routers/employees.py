from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session
from typing import List, Optional, Dict, Any
from io import BytesIO, StringIO
import csv
import re

from app.database import get_db
from app.models import User
from app.schemas import UserOut, UserCreate, UserUpdate
from app.routers.auth import get_current_user
from app.auth import get_password_hash

router = APIRouter()


def require_admin_hr(current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "hr"]:
        raise HTTPException(status_code=403, detail="Admin or HR access required")
    return current_user


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_header(header: str) -> str:
    h = clean_text(header).lower()
    h = h.replace(" ", "_").replace("-", "_")
    h = re.sub(r"[^a-z0-9_]", "", h)

    aliases = {
        "employeeid": "employee_id",
        "emp_id": "employee_id",
        "empid": "employee_id",
        "employee_code": "employee_id",
        "fullname": "name",
        "full_name": "name",
        "employee_name": "name",
        "mail": "email",
        "email_id": "email",
        "mobile": "phone",
        "phone_number": "phone",
        "contact": "phone",
        "job_title": "designation",
        "position": "designation",
        "dept": "department",
        "workmode": "work_mode",
        "work_location": "work_mode",
    }

    return aliases.get(h, h)


def generate_employee_id(db: Session) -> str:
    count = db.query(User).count() + 1
    return f"EMP{str(count).zfill(3)}"


def create_or_skip_employee(row: Dict[str, Any], db: Session) -> Dict[str, Any]:
    name = clean_text(row.get("name"))
    email = clean_text(row.get("email")).lower()

    if not name or not email:
        return {
            "status": "skipped",
            "reason": "Missing name or email",
            "row": row,
        }

    existing = db.query(User).filter(User.email == email).first()

    if existing:
        return {
            "status": "skipped",
            "reason": "Email already exists",
            "email": email,
        }

    employee_id = clean_text(row.get("employee_id")) or generate_employee_id(db)

    existing_emp_id = db.query(User).filter(User.employee_id == employee_id).first()
    if existing_emp_id:
        employee_id = generate_employee_id(db)

    password = clean_text(row.get("password")) or "password123"

    user = User(
        email=email,
        password_hash=get_password_hash(password),
        name=name,
        role=clean_text(row.get("role")) or "employee",
        department=clean_text(row.get("department")) or "General",
        designation=clean_text(row.get("designation")) or "Employee",
        phone=clean_text(row.get("phone")),
        employee_id=employee_id,
        status=clean_text(row.get("status")) or "active",
        work_mode=clean_text(row.get("work_mode")) or "onsite",
    )

    db.add(user)
    db.flush()

    return {
        "status": "created",
        "name": user.name,
        "email": user.email,
        "employee_id": user.employee_id,
    }


def rows_from_csv_bytes(content: bytes) -> List[Dict[str, Any]]:
    text = content.decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(StringIO(text))
    rows = []

    for raw_row in reader:
        row = {}
        for key, value in raw_row.items():
            row[normalize_header(key)] = clean_text(value)
        rows.append(row)

    return rows


def rows_from_excel_bytes(content: bytes) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is not installed. Run: pip install openpyxl"
        )

    workbook = load_workbook(BytesIO(content), data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [normalize_header(str(cell or "")) for cell in rows[0]]
    parsed_rows = []

    for values in rows[1:]:
        row = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            row[header] = clean_text(values[index] if index < len(values) else "")
        if any(row.values()):
            parsed_rows.append(row)

    return parsed_rows


def extract_text_from_pdf(content: bytes) -> str:
    try:
        from pypdf import PdfReader
    except Exception:
        raise HTTPException(
            status_code=500,
            detail="pypdf is not installed. Run: pip install pypdf"
        )

    reader = PdfReader(BytesIO(content))
    text_parts = []

    for page in reader.pages:
        text_parts.append(page.extract_text() or "")

    return "\n".join(text_parts)


def extract_text_from_docx(content: bytes) -> str:
    try:
        from docx import Document as DocxDocument
    except Exception:
        raise HTTPException(
            status_code=500,
            detail="python-docx is not installed. Run: pip install python-docx"
        )

    document = DocxDocument(BytesIO(content))
    return "\n".join([p.text for p in document.paragraphs])


def rows_from_loose_text(text: str) -> List[Dict[str, Any]]:
    """
    Supports simple text/PDF/DOCX formats like:
    Name, Email, Department, Designation, Phone
    Rahul Kumar, rahul@email.com, Engineering, Developer, 9876543210

    Also supports lines like:
    Name: Rahul Kumar | Email: rahul@email.com | Department: Engineering | Designation: Developer
    """
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    parsed_rows = []

    # Try CSV-like text first.
    if len(lines) >= 2 and "," in lines[0]:
        reader = csv.DictReader(StringIO("\n".join(lines)))
        for raw_row in reader:
            row = {}
            for key, value in raw_row.items():
                row[normalize_header(key)] = clean_text(value)
            if any(row.values()):
                parsed_rows.append(row)

        if parsed_rows:
            return parsed_rows

    # Try key-value format.
    current = {}

    for line in lines:
        parts = re.split(r"\||,", line)
        line_data = {}

        for part in parts:
            if ":" in part:
                key, value = part.split(":", 1)
                line_data[normalize_header(key)] = clean_text(value)

        if line_data.get("name") or line_data.get("email"):
            if current:
                parsed_rows.append(current)
            current = line_data
        else:
            current.update(line_data)

    if current:
        parsed_rows.append(current)

    # Last fallback: detect email and use nearby text as name.
    if not parsed_rows:
        for line in lines:
            email_match = re.search(r"[\w\.-]+@[\w\.-]+\.\w+", line)
            if email_match:
                email = email_match.group(0)
                before_email = line[:email_match.start()].strip(" -,:|")
                parsed_rows.append({
                    "name": before_email or email.split("@")[0].replace(".", " ").title(),
                    "email": email,
                    "department": "General",
                    "designation": "Employee",
                    "role": "employee",
                    "work_mode": "onsite",
                    "status": "active",
                })

    return parsed_rows


def parse_employee_file(filename: str, content: bytes) -> List[Dict[str, Any]]:
    lower = filename.lower()

    if lower.endswith(".csv"):
        return rows_from_csv_bytes(content)

    if lower.endswith(".xlsx") or lower.endswith(".xls"):
        return rows_from_excel_bytes(content)

    if lower.endswith(".pdf"):
        text = extract_text_from_pdf(content)
        return rows_from_loose_text(text)

    if lower.endswith(".docx"):
        text = extract_text_from_docx(content)
        return rows_from_loose_text(text)

    if lower.endswith(".txt"):
        text = content.decode("utf-8-sig", errors="ignore")
        return rows_from_loose_text(text)

    raise HTTPException(
        status_code=400,
        detail="Unsupported file type. Upload CSV, XLSX, PDF, DOCX, or TXT."
    )


@router.get("/", response_model=List[UserOut])
def get_employees(
    skip: int = 0,
    limit: int = 100,
    search: Optional[str] = Query(None),
    department: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_hr)
):
    query = db.query(User)

    if search:
        pattern = f"%{search}%"
        query = query.filter(
            (User.name.ilike(pattern)) |
            (User.email.ilike(pattern)) |
            (User.employee_id.ilike(pattern))
        )

    if department:
        query = query.filter(User.department == department)

    if status:
        query = query.filter(User.status == status)

    return query.offset(skip).limit(limit).all()


@router.post("/", response_model=UserOut)
def create_employee(
    user: UserCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_hr)
):
    db_user = db.query(User).filter(User.email == user.email).first()

    if db_user:
        raise HTTPException(status_code=400, detail="Email already registered")

    emp_count = db.query(User).count()
    employee_id = user.employee_id or f"EMP{str(emp_count + 1).zfill(3)}"

    hashed_password = get_password_hash(user.password)

    db_user = User(
        email=user.email,
        password_hash=hashed_password,
        name=user.name,
        role=user.role or "employee",
        department=user.department or "General",
        designation=user.designation or "Employee",
        phone=user.phone or "",
        employee_id=employee_id,
        status=user.status or "active",
        work_mode=user.work_mode or "onsite"
    )

    db.add(db_user)
    db.commit()
    db.refresh(db_user)

    return db_user


@router.post("/bulk-import")
def bulk_import_employees(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_hr)
):
    content = file.file.read()

    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    rows = parse_employee_file(file.filename or "", content)

    if not rows:
        raise HTTPException(
            status_code=400,
            detail="Could not find employee rows in the uploaded file"
        )

    results = []
    created_count = 0
    skipped_count = 0

    try:
        for row in rows:
            result = create_or_skip_employee(row, db)
            results.append(result)

            if result["status"] == "created":
                created_count += 1
            else:
                skipped_count += 1

        db.commit()

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Bulk import failed: {str(e)}")

    return {
        "message": "Bulk import completed",
        "created": created_count,
        "skipped": skipped_count,
        "total_rows_detected": len(rows),
        "results": results,
        "default_password": "password123"
    }


@router.get("/{user_id}", response_model=UserOut)
def get_employee(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if current_user.role not in ["admin", "hr"] and current_user.id != user_id:
        raise HTTPException(status_code=403, detail="Access denied")

    user = db.query(User).filter(User.id == user_id).first()

    if not user:
        raise HTTPException(status_code=404, detail="Employee not found")

    return user


@router.patch("/{user_id}", response_model=UserOut)
def update_employee(
    user_id: int,
    user_update: UserUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_hr)
):
    user = db.query(User).filter(User.id == user_id).first()

    if not user:
        raise HTTPException(status_code=404, detail="Employee not found")

    update_data = user_update.dict(exclude_unset=True)

    for field, value in update_data.items():
        setattr(user, field, value)

    db.commit()
    db.refresh(user)

    return user


@router.delete("/{user_id}")
def delete_employee(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_hr)
):
    user = db.query(User).filter(User.id == user_id).first()

    if not user:
        raise HTTPException(status_code=404, detail="Employee not found")

    if user.role == "admin" and current_user.id != user_id:
        raise HTTPException(status_code=400, detail="Cannot delete admin")

    db.delete(user)
    db.commit()

    return {"message": "Employee deleted successfully"}